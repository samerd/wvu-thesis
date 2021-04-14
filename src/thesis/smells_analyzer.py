'''
Created on Jun 7, 2020

@author: samerd
'''
from __future__ import (absolute_import, division)

import os
import scipy.stats

import numpy
from openpyxl import Workbook
from openpyxl.chart import BarChart3D, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.fills import PatternFill

from thesis import get_dataset_dir
from thesis.designite_projects import DesigniteProjects
from thesis.metrics_mgr import MetricsMgr
from thesis.package_efforts import PackageEffortsAnalyzer
from thesis.release_details_mgr import ReleaseDetailsMgr
from thesis.smells_parsers import ArchSmellsParserEx, DesignSmellsParser, \
    PackageInfo


def _add_chart(work_sheet, cols, rows, title, pos):
    chart = BarChart3D()
    chart.title = title
    for col in cols:
        data = Reference(
            work_sheet, min_col=col, min_row=1, max_row=rows)
        chart.add_data(data, titles_from_data=True)
    categories = Reference(
        work_sheet, min_col=1, min_row=2, max_row=rows)
    chart.set_categories(categories)
    chart.shape = 'box'
    work_sheet.add_chart(chart, "B%s" % pos)


class ProjectInfo:

    # pylint: disable=too-many-instance-attributes
    def __init__(self, proj, release, next_release):
        self.proj = proj
        self.release = release
        self.next_release = next_release
        self.classes = None
        self.packages = None
        self.loc = None
        self.rank_cc = None
        self.score_cc = None
        self.rank_p = None
        self.score_p = None
        self.effort = None
        self.arch_smells = 0
        self.design_smells = 0

    @classmethod
    def get_headers(cls):
        return ("Project", "Release", "Next Release", "Packages", "Classes",
                "Total LOC", "Rank CC", "Rank p-value", "Score CC",
                "Score p-value", "Effort", "Arch-Smells", "Design-Smells")

    def to_tuple(self):
        return (self.proj, self.release, self.next_release, self.packages,
                self.classes, self.loc, self.rank_cc, self.rank_p,
                self.score_cc, self.score_p, self.effort, self.arch_smells,
                self.design_smells)


class CorrelationStats:
    # pylint: disable=too-many-instance-attributes

    def __init__(self):
        self.rank_arr = list()
        self.effort_rank_arr = list()
        self.score_arr = list()
        self.effort_arr = list()
        self.rank_cc = 0
        self.rank_p = 0
        self.score_cc = 0
        self.score_p = 0

    def add(self, score, rank, effort, effort_rank):
        self.effort_arr.append(effort)
        self.effort_rank_arr.append(effort_rank)
        self.rank_arr.append(rank)
        self.score_arr.append(score)

    def calc(self):
        self.rank_cc, self.rank_p = scipy.stats.spearmanr(
            self.rank_arr, self.effort_rank_arr)
        self.score_cc, self.score_p = scipy.stats.spearmanr(
            self.score_arr, self.effort_arr)


class ProjectSmellsAnalyzer:
    # pylint: disable=too-many-instance-attributes,too-many-arguments

    def __init__(self, proj, release, next_release, designite_path,
                 out_folder, pkg_analyzer, release_mgr):
        self.proj_info = ProjectInfo(proj, release, next_release)
        self.designite_path = designite_path
        self.packages_info = dict()
        file_path = os.path.join(self.designite_path, "TypeMetrics.csv")
        self._metrics = MetricsMgr(file_path)
        self.proj_info.classes = self._metrics.get_classes_count()
        self.proj_info.packages = self._metrics.get_pkg_count()
        self.proj_info.loc = self._metrics.get_total_loc()
        self.out_folder = out_folder
        self.pkg_analyzer = pkg_analyzer
        self.release_mgr = release_mgr
        self.work_book = Workbook()
        self.arch_sheet = self.work_book.active
        self.arch_sheet.title = "ArchSmells Ranks"
        self.design_sheet = self.work_book.create_sheet("DesignSmells Ranks")
        self.total_sheet = self.work_book.create_sheet("TotalSmells Ranks")
        self.sig_sheet = self.work_book.create_sheet(
            "Arch-Design-Smells Ranks")

    @classmethod
    def _add_correlation(cls, work_sheet, corr_stats):
        corr_stats.calc()
        work_sheet.append(("",))
        work_sheet.append(("", "Score", "Rank"))
        work_sheet.append(("Spearman CC", "%.2f" % corr_stats.score_cc,
                           "%.2f" % corr_stats.rank_cc))
        work_sheet.append(("p-value", "%.3f" % corr_stats.score_p,
                           "%.3f" % corr_stats.rank_p))

    @classmethod
    def _add_chart(cls, work_sheet, cols, rows, title):
        chart = BarChart3D()
        chart.title = title
        for col in cols:
            data = Reference(
                work_sheet, min_col=col, min_row=1, max_row=rows)
            chart.add_data(data, titles_from_data=True)
        categories = Reference(
            work_sheet, min_col=1, min_row=2, max_row=rows)
        chart.set_categories(categories)
        chart.shape = 'box'
        work_sheet.add_chart(chart, "B%s" % (rows + 6))

    def _set_pkg_info(self, pkg):
        self.packages_info[pkg] = PackageInfo(pkg)
        self.packages_info[pkg].class_count = \
            self._metrics.get_pkg_classes(pkg)
        self.packages_info[pkg].loc = self._metrics.get_pkg_loc(pkg)
        self.packages_info[pkg].metrics = self._metrics.get_pkg_metrics(pkg)
        pkg_efforts = self.pkg_analyzer.get_pkg_efforts(
            self.proj_info.proj, self.proj_info.next_release, pkg)
        if pkg_efforts:
            self.packages_info[pkg].lines_added = pkg_efforts.lines_added

    def analyze_smells(self):
        file_path = os.path.join(self.designite_path, "ArchitectureSmells.csv")
        parser = ArchSmellsParserEx(file_path)
        self.proj_info.arch_smells = 0
        if "<All packages>" in parser.data:
            parser.data.pop("<All packages>")
        for pkg, pkg_data in parser.data.items():
            if pkg not in self.packages_info:
                self._set_pkg_info(pkg)

            pkg_info = self.packages_info[pkg]

            pkg_val = 0
            for _, val in pkg_data.items():
                pkg_val += val
            pkg_info.arch_count = pkg_val
            self.proj_info.arch_smells += pkg_info.arch_count
            pkg_info.arch_score = pkg_val / parser.get_smells_count()
            pkg_smells = parser.get_pkg_smells(pkg)
            if pkg_smells:
                pkg_info.smell_details = pkg_smells

        file_path = os.path.join(self.designite_path, "DesignSmells.csv")
        parser = DesignSmellsParser(file_path)
        for pkg, pkg_data in parser.data.items():
            if pkg not in self.packages_info:
                self._set_pkg_info(pkg)

        rows = 0
        columns = ("Package", "ArchSmells", "ArchScore", "Rank", "Effort",
                   "EffortRank")
        rows += 1
        self.arch_sheet.append(columns)
        sorted_pkg = sorted(
            self.packages_info.items(), key=lambda x: x[1].arch_score,
            reverse=True)
        corr_stats = CorrelationStats()
        rank = 0
        pkgs_ph_ranks = self.pkg_analyzer.get_pkgs_ranks(
            self.proj_info.proj, self.proj_info.next_release,
            self.packages_info)
        count = len(sorted_pkg)
        for pkg, pkg_info in sorted_pkg:
            rank += 1
            pkg_info.set_smells_rank(rank, count)
            effort, effort_rank = pkgs_ph_ranks.get(
                pkg, [0, count])
            pkg_info.set_effort_rank(effort_rank, count)
            corr_stats.add(pkg_info.arch_score, rank, effort, effort_rank)
            self.arch_sheet.append((
                pkg, pkg_info.arch_count, pkg_info.arch_score, rank, effort,
                effort_rank))
            rows += 1
        self._add_correlation(self.arch_sheet, corr_stats)
        self.proj_info.rank_cc = corr_stats.rank_cc
        self.proj_info.rank_p = corr_stats.rank_p
        self.proj_info.score_cc = corr_stats.score_cc
        self.proj_info.score_p = corr_stats.score_p
        release_info = self.release_mgr.get_release_info(
            self.proj_info.proj, self.proj_info.next_release)
        if release_info:
            self.proj_info.effort = release_info.effort

        rank_col = columns.index("Rank") + 1
        effor_rank_col = columns.index("EffortRank") + 1

        cols = (rank_col, effor_rank_col)
        title = "Estimated Effort vs. Actual Effort Based on Arch Smells"
        _add_chart(self.arch_sheet, cols, rows, title, rows + 6)

    def save(self):
        fname = os.path.join(self.out_folder, "designite_analysis.xlsx")
        self.work_book.save(fname)


class DesigniteAnalyzer:

    def __init__(self):
        self._pkg_analyzer = PackageEffortsAnalyzer(
            DesigniteProjects.PROJECTS)
        self._release_mgr = ReleaseDetailsMgr()

    def load_data(self):
        ref_miner_dir = os.path.join(get_dataset_dir(), "ref_miner")
        print("Loading packages efforts...")
        self._pkg_analyzer.load(
            os.path.join(ref_miner_dir, "ref_package_ex.csv"))
        print("Done")
        self._release_mgr.load(
            os.path.join(ref_miner_dir, "ref_release_summ.csv"))

    def analyze_projects(self):
        work_book = Workbook()
        ws1 = work_book.active
        ws1.title = "All Projects"
        ws2 = work_book.create_sheet("Filtered Projects")
        parent_designite_path = os.path.join(get_dataset_dir(), "designite")
        out_folder = os.path.join(get_dataset_dir(), "designite_analysis")
        if not os.path.isdir(out_folder):
            os.mkdir(out_folder)
        ws1.append(ProjectInfo.get_headers())
        ws2.append(ProjectInfo.get_headers())
        cc_list = list()
        classes_list = list()
        pkg_list = list()
        arch_smells_list = list()
        design_smells_list = list()

        fp3 = open(os.path.join(out_folder, "pkg_efforts_3.arff"), "w")
        fp3_nd = open(os.path.join(out_folder, "pkg_efforts_3_nd.arff"), "w")
        fp5 = open(os.path.join(out_folder, "pkg_efforts_5.arff"), "w")
        fp5_nd = open(os.path.join(out_folder, "pkg_efforts_5_nd.arff"), "w")
        fp_ext = open(os.path.join(out_folder, "pkg_efforts_ext.csv"), "w")

        fp3.write(PackageInfo.get_arff_header("Efforts-3-Levels"))
        fp_ext.write(PackageInfo.get_extended_header())
        fp3_nd.write(PackageInfo.get_arff_header(
            "Efforts-3-Levels-Normal-Distribution"))
        fp5.write(PackageInfo.get_arff_header("Efforts-5-Levels"))
        fp5_nd.write(PackageInfo.get_arff_header(
            "Efforts-5-Levels-Normal-Distribution"))
        for proj, releases in DesigniteProjects.PROJECTS.items():
            for release, next_release in releases.items():
                print("Analyzing project:", proj, "release:", release)
                proj_path = "%s-%s" % (proj, release)
                designite_path = os.path.join(
                    parent_designite_path, proj, release)
                proj_out_path = os.path.join(
                    out_folder, proj_path)
                if not os.path.isdir(proj_out_path):
                    os.mkdir(proj_out_path)
                proj_analyzer = ProjectSmellsAnalyzer(
                    proj, release, next_release, designite_path,
                    proj_out_path, self._pkg_analyzer, self._release_mgr)

                proj_analyzer.analyze_smells()
                for pkg, pkg_info in proj_analyzer.packages_info.items():
                    if pkg == "<All packages>":
                        continue
                    fp3.write(pkg_info.get_arff_3())
                    fp3_nd.write(pkg_info.get_arff_3_nd())
                    fp5.write(pkg_info.get_arff_5())
                    fp5_nd.write(pkg_info.get_arff_5_nd())
                    fp_ext.write(pkg_info.get_extended_data(proj, release, pkg))
                proj_analyzer.save()
                if proj_analyzer.proj_info.rank_p < 0.05:
                    ws2.append(proj_analyzer.proj_info.to_tuple())
                    cc_list.append(proj_analyzer.proj_info.rank_cc)
                    design_smells_list.append(
                        proj_analyzer.proj_info.design_smells)
                    arch_smells_list.append(
                        proj_analyzer.proj_info.arch_smells)
                    classes_list.append(proj_analyzer.proj_info.classes)
                    pkg_list.append(proj_analyzer.proj_info.packages)
                ws1.append(proj_analyzer.proj_info.to_tuple())
        fp3.close()
        fp3_nd.close()
        fp5.close()
        fp5_nd.close()
        fp_ext.close()
        ws2.append(("",))
        ws2.append(("Median", numpy.median(cc_list)))
        ws2.append(("Mean", numpy.mean(cc_list)))
        ws2.append(("Stdev", numpy.std(cc_list)))
        ws2.append(("Min", numpy.min(cc_list)))
        ws2.append(("Max", numpy.max(cc_list)))
        data_range = "H2:H%d" % (len(cc_list) + 1)
        red_fill = PatternFill(start_color='EE1111',
                               end_color='EE1111',
                               fill_type='solid')
        ws1.conditional_formatting.add(
            data_range, CellIsRule(operator='greaterThan', formula=[0.05],
                                   stopIfTrue=False, fill=red_fill))
        filename = os.path.join(out_folder, "projects_analysis.xlsx")
        col_pos = ProjectInfo.get_headers().index("Rank CC") + 1
        _add_chart(ws2, (col_pos,), len(cc_list) + 1, "Spearman Correlation",
                   len(cc_list) + 10)
        self._analyze_correlation(
            cc_list, classes_list, pkg_list, arch_smells_list,
            design_smells_list)
        work_book.save(filename)

    @classmethod
    def _analyze_correlation(cls, cc_list, classes_list, pkg_list,
                             arch_smells_list, design_smells_list):
        corr, pvalue = scipy.stats.spearmanr(cc_list, classes_list)
        print(corr, pvalue)
        corr, pvalue = scipy.stats.spearmanr(cc_list, pkg_list)
        print(corr, pvalue)
        corr, pvalue = scipy.stats.spearmanr(cc_list, arch_smells_list)
        print(corr, pvalue)
        corr, pvalue = scipy.stats.spearmanr(cc_list, design_smells_list)
        print(corr, pvalue)


if __name__ == "__main__":
    analyzer = DesigniteAnalyzer()
    analyzer.load_data()
    analyzer.analyze_projects()
